'''
########################################################
# Name Project: Attendance List                        #
# Developer Name: Mário Wessen                         #
# Contact: mrfwn@cin.ufpe.br                           #
# Date last Modify:  24/06/2019                        #
# Description File:This is Class for manipuled         #
# Contacts. Methods for automation the system          #
########################################################
'''
import os
from openpyxl import load_workbook
from app.services.database_service import Database
import qrcode
from qrcode.image.pure import PymagingImage
from PIL import Image, ImageFont, ImageDraw
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
import base64
import glob


class Contact:
    UPLOAD_FOLDER = "./app/uploads/"
    BASE_FOLDER = "./app/uploads/imgbase/"
    INVITE_FOLDER = "./app/uploads/invites/"
    DBNAME = "contacts"
    db = None
    agencys = []

    def __init__(self):
        self.db = Database(self.DBNAME)

    # Methos for load List the contacts. OBS: Is necessary especific format in file.
    # [name,agency,email,telefone]
    # The table in file don't have Title
    def loadList(self, fileName):
        fileName = self.UPLOAD_FOLDER + fileName
        #listInstance = load_workbook(filename=fileName, read_only=True).active
        listInstance = load_workbook(filename=fileName, read_only=True)
        line = 1
        contacts = []
        listRegisted = self.db.get_datas()
        while (listInstance.active.cell(row=line, column=1).value != None):
            contacts.append(
                {
                    "name": listInstance.active.cell(row=line, column=1).value,
                    "agency": listInstance.active.cell(row=line, column=2).value,
                    "email1": listInstance.active.cell(row=line, column=3).value,
                    "email2": listInstance.active.cell(row=line, column=4).value != None if listInstance.active.cell(row=line, column=4).value else "",
                    "tel1": listInstance.active.cell(row=line, column=5).value,
                    "tel2": listInstance.active.cell(row=line, column=6).value != None if listInstance.active.cell(row=line, column=6).value else "",
                    "sexo": listInstance.active.cell(row=line, column=7).value,
                    "event": listInstance.active.cell(row=line, column=8).value,
                    "office": listInstance.active.cell(row=line, column=9).value,
                    "status": 0,
                    "regtype": 1,
                    "invited": 0
                }
            )
            line += 1
        listInstance.close()
        if listRegisted :
            for key,val in listRegisted.items():
                for index in contacts:
                    if index['email1'] == val['email1']:
                        contacts.remove(index)
        self.db.push_list_data(contacts)

    # This Method Generate Invites, is necessary set parameters for positioning images QR,base and Name
    def generateInvit(self):
        areaQR = (220, 750)
        areaName = (240, 700)
        font = ImageFont.truetype("arial.ttf", 45)
        contacts = self.db.get_datas()
        for key in contacts:
            qrcodeImg = qrcode.make(key)
            baseImg = Image.open(self.BASE_FOLDER+'base_conv.png')
            colagem = ImageDraw.Draw(baseImg)
            colagem.text(areaName, contacts[key]['name'], font=font, fill='black')
            baseImg.paste(qrcodeImg, areaQR)
            baseImg.save(self.INVITE_FOLDER+contacts[key]['email1']+".png")

# This Method Send Email for any contacts OBS:is necessary generate Invite before

    def content(self, nome):
        return """
        <html>
        <head>
        <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
        <title>Teste envio</title>
        <style type="text/css">
            a {color: #d80a3e;}
        body, #header h1, #header h2, p {margin: 0; padding: 0;}

        #main {border: 1px solid #cfcece;}
        img {display: block;}
        #top-message p, #bottom p {color: #3f4042; font-size: 12px; font-family: Arial, Helvetica, sans-serif; }
        #header h4 {color: #ffffff !important; font-family: "Lucida Grande", sans-serif; font-size: 24px; margin-bottom: 0!important; padding-bottom: 0; }
        #header p {color: #ffffff !important; font-family: "Lucida Grande", "Lucida Sans", "Lucida Sans Unicode", sans-serif; font-size: 12px;  }
        h5 {margin: 0 0 0.8em 0;}
            h5 {font-size: 18px; color: #444444 !important; font-family: Arial, Helvetica, sans-serif; }
        p {font-size: 12px; color: #444444 !important; font-family: "Lucida Grande", "Lucida Sans", "Lucida Sans Unicode", sans-serif; line-height: 1.5;}
        </style>
        </head>
        <body><h4>""" + str(nome) + """</h4>
            <img src="cid:myimage"/>
        </body>
        </html> 
        """
    def getListUser(self):
        contacts = self.db.get_datas()
        objList ={
                "total": 0,
                "totalNotFiltered": 0,
                "rows": []
            }
        if contacts:
            objList ={
                "total": len(contacts),
                "totalNotFiltered": len(contacts),
                "rows": []
            }
            i = 0
            for key,val in contacts.items():
                objList['rows'].append({
                'id': val['name'],
                'name':val['email1'],
                'price':val['office'],
                'index': i
                })
                i+=1
        return objList
        

    # This Method Send Email for any contacts OBS:is necessary generate Invite before
    def sendEmail(self):
        areaQR = (160, 650)
        areaName = (160, 1080)
        font = ImageFont.truetype("./app/static/webfonts/Hurufo-&-Numero.ttf", 80)
        contacts = self.db.get_datas()
        for key in contacts:
            if contacts[key]['invited'] != 1:
                self.db.update_data(key,{"invited":1})
                fromEmail = 'eventos.mktpe@tvglobo.com.br'
                password = 'JCw35zwxxw5a239zwFMAaagbgTU78f'
                qr = qrcode.QRCode(box_size=16,border=1)
                qr.add_data(key)
                qr.make(key)
                qrcodeImg = qr.make_image()
                baseImg = Image.open(self.BASE_FOLDER+'convite_QRCODE.png')
                colagem = ImageDraw.Draw(baseImg)
                colagem.text(areaName, contacts[key]['name'], font=font, fill='black')
                baseImg.paste(qrcodeImg, areaQR)
                baseImg.save(self.INVITE_FOLDER+contacts[key]['email1']+".png")
                email_content = self.content("")
                toEmail = contacts[key]['email1']
                subject = 'Chegou o seu QR Code! | PDM 2019'
                msg = MIMEMultipart()
                msg['From'] = fromEmail
                msg['To'] = toEmail
                msg['Subject'] = subject
                img_data = open(self.INVITE_FOLDER + toEmail + '.png', 'rb').read()
                body = MIMEText(email_content, _subtype='html')
                msg.attach(body)
                img = MIMEImage(img_data, 'png')
                img.add_header('Content-Id', '<myimage>')
                msg.attach(img)
                server = smtplib.SMTP("smtp.office365.com", 587)
                server.ehlo()
                server.starttls()
                server.ehlo()
                server.login(fromEmail, password)
                server.sendmail(fromEmail, toEmail, msg.as_string())
                server.close()

    def clearList(self):
        fileList = glob.glob(self.INVITE_FOLDER+"*.png")
        for filePath in fileList:
            try:
                os.remove(filePath)
            except:
                print("Error while deleting file: ", filePath)
        self.db.remove_all_data()

    def countPresence(self):
        presence = {
            'presence': 0,
            'fault': 0
        }
        if self.db.get_datas():
            contacts = self.db.get_datas()
            for key in contacts:
                if contacts[key]['status'] == 0:
                    presence['fault'] += 1
                else:
                    presence['presence'] += 1
        return presence

    def countAgence(self):
        agencyCount = []
        if self.db.get_datas():
            contacts = self.db.get_datas()
            self.agencys = []
            agencysAll = []
            for key in contacts:
                agencysAll.append(contacts[key]['agency'])
                if contacts[key]['agency'] not in self.agencys:
                    self.agencys.append(contacts[key]['agency'])
            for agency in self.agencys:
                agencyCount.append(agencysAll.count(agency))
        return agencyCount
